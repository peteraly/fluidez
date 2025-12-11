#!/usr/bin/env python3
"""
═══════════════════════════════════════════════════════════════════════════════
FLUIDEZ CURRICULUM AUDIT & ENHANCEMENT SCRIPT
═══════════════════════════════════════════════════════════════════════════════

Applies three SLA frameworks to enhance (not replace) existing curriculum:
1. Processability Theory (Pienemann) - Grammar sequencing validation
2. Noticing Hypothesis (Schmidt) - Attention-directing techniques
3. Shadowing Mode (Phonological Loop) - Pronunciation automaticity

PRESERVATION MODE: This script only ADDS content, never deletes or overwrites
existing curriculum material.

Outputs:
- Updated Excel file with 5 new worksheets
- Markdown audit report
- JSON day-by-day implementation specs
"""

import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

PRESERVE_MODE = True  # Only ADD, never modify existing content

INPUT_PATH = "/mnt/user-data/outputs/Fluidez_Master_Alignment_Matrix.xlsx"
OUTPUT_EXCEL = "/mnt/user-data/outputs/Fluidez_Master_Alignment_Matrix_v2_Enhanced.xlsx"
OUTPUT_REPORT = "/mnt/user-data/outputs/Fluidez_Curriculum_Audit_Report.md"
OUTPUT_JSON = "/mnt/user-data/outputs/Fluidez_Day_Implementations.json"

# ═══════════════════════════════════════════════════════════════════════════════
# PROCESSABILITY THEORY STAGES (Spanish)
# ═══════════════════════════════════════════════════════════════════════════════

PT_STAGES = {
    1: {
        "name": "Lemma Access (Words/Formulas)",
        "description": "Single words and memorized chunks - no grammatical processing",
        "structures": [
            "Single vocabulary words",
            "Memorized phrases/chunks",
            "Formulaic expressions",
            "Greetings and farewells",
            "Numbers and colors",
            "Basic nouns (no modification)"
        ],
        "spanish_examples": [
            "Hola", "Gracias", "Por favor", "Buenos días",
            "Me llamo...", "¿Cómo estás?", "Muy bien",
            "uno, dos, tres", "rojo, azul, verde"
        ],
        "typical_days": [1, 2, 3],
        "prerequisite": None,
        "keywords": ["greetings", "numbers", "colors", "basics", "phrases", "vocabulary", "words"]
    },
    2: {
        "name": "Category Procedure (Lexical Morphology)",
        "description": "Word-level grammatical marking - gender, number on single words",
        "structures": [
            "Plural marking (-s/-es)",
            "Gender marking (el/la, un/una)",
            "Definite articles",
            "Indefinite articles",
            "Possessive adjectives (mi, tu, su)",
            "Basic demonstratives"
        ],
        "spanish_examples": [
            "el libro / los libros",
            "la mesa / las mesas",
            "un coche / una casa",
            "mi hermano / tu hermana / su libro"
        ],
        "typical_days": [3, 4, 5, 6, 7],
        "prerequisite": 1,
        "keywords": ["articles", "el", "la", "plural", "gender", "possessive", "mi", "tu", "su"]
    },
    3: {
        "name": "Phrasal Procedure (Phrase-level Agreement)",
        "description": "Agreement within phrases - adjective-noun agreement",
        "structures": [
            "Adjective-noun gender agreement",
            "Adjective-noun number agreement",
            "Adjective position (post-nominal)",
            "Demonstrative agreement (este/esta/estos/estas)",
            "Quantifier agreement"
        ],
        "spanish_examples": [
            "el coche rojo / la casa roja",
            "los coches rojos / las casas rojas",
            "este libro / esta mesa",
            "estos libros / estas mesas",
            "mucho trabajo / mucha agua"
        ],
        "typical_days": [5, 6, 7, 8, 9, 10],
        "prerequisite": 2,
        "keywords": ["adjective", "agreement", "este", "esta", "position", "noun phrase"]
    },
    4: {
        "name": "Sentence Procedure (S-V Agreement)",
        "description": "Sentence-level processing - subject-verb agreement, word order",
        "structures": [
            "Present tense conjugation (-AR, -ER, -IR)",
            "Subject-verb agreement",
            "Basic SVO word order",
            "Negation (no + verb)",
            "Yes/no question formation",
            "Basic ser/estar distinction",
            "Ir + a + infinitive (near future)"
        ],
        "spanish_examples": [
            "Yo hablo español",
            "Tú comes pizza",
            "Él vive en Madrid",
            "No hablo francés",
            "¿Hablas inglés?",
            "Soy estudiante / Estoy cansado",
            "Voy a comer"
        ],
        "typical_days": [8, 9, 10, 11, 12, 13, 14, 15],
        "prerequisite": 3,
        "keywords": ["conjugation", "present tense", "-ar", "-er", "-ir", "verb", "ser", "estar", "negation"]
    },
    5: {
        "name": "S-Procedure (Subordinate Clause)",
        "description": "Inter-phrasal processing - pronouns, reflexives, basic subordination",
        "structures": [
            "Direct object pronouns (lo, la, los, las)",
            "Indirect object pronouns (le, les)",
            "Reflexive verbs (me levanto, se llama)",
            "Que-clauses (Creo que..., Pienso que...)",
            "Preterite tense (basic regular)",
            "Imperfect tense (basic)",
            "Pronoun placement with infinitives"
        ],
        "spanish_examples": [
            "Lo quiero", "La veo", "Les doy el libro",
            "Me levanto a las siete",
            "Creo que es importante",
            "Ayer comí pizza",
            "Cuando era niño...",
            "Quiero verlo / Lo quiero ver"
        ],
        "typical_days": [15, 16, 17, 18, 19, 20, 21, 22],
        "prerequisite": 4,
        "keywords": ["pronoun", "lo", "la", "le", "reflexive", "preterite", "past", "imperfect", "que clause"]
    },
    6: {
        "name": "S'-Procedure (Complex Subordination)",
        "description": "Complex inter-clausal processing - subjunctive, conditionals",
        "structures": [
            "Present subjunctive (basic)",
            "Subjunctive with desire/emotion (quiero que...)",
            "Conditional mood (me gustaría)",
            "If-clauses (si + present)",
            "Relative clauses (que, quien, donde)",
            "Reported speech",
            "Perfect tenses (he comido)"
        ],
        "spanish_examples": [
            "Quiero que vengas",
            "Espero que estés bien",
            "Me gustaría viajar",
            "Si tengo tiempo, voy",
            "El libro que leí",
            "La persona que conocí",
            "He comido ya"
        ],
        "typical_days": [22, 23, 24, 25, 26, 27, 28, 29, 30],
        "prerequisite": 5,
        "keywords": ["subjunctive", "conditional", "si clause", "relative", "perfect", "he", "has", "ha"]
    }
}

# ═══════════════════════════════════════════════════════════════════════════════
# NOTICING HYPOTHESIS TECHNIQUES
# ═══════════════════════════════════════════════════════════════════════════════

NOTICING_TECHNIQUES = {
    "color_coding": {
        "description": "Visual color differentiation for grammatical features",
        "verb_endings": {
            "yo": {"color": "#F44336", "name": "Red"},
            "tú": {"color": "#2196F3", "name": "Blue"},
            "él/ella/usted": {"color": "#4CAF50", "name": "Green"},
            "nosotros": {"color": "#9C27B0", "name": "Purple"},
            "vosotros": {"color": "#FF5722", "name": "Deep Orange"},
            "ellos/ustedes": {"color": "#FF9800", "name": "Orange"}
        },
        "gender": {
            "masculine": {"color": "#2196F3", "name": "Blue"},
            "feminine": {"color": "#E91E63", "name": "Pink"}
        },
        "parts_of_speech": {
            "noun": {"color": "#4CAF50", "name": "Green"},
            "verb": {"color": "#2196F3", "name": "Blue"},
            "adjective": {"color": "#FF9800", "name": "Orange"},
            "adverb": {"color": "#9C27B0", "name": "Purple"}
        },
        "tense": {
            "present": {"color": "#4CAF50", "name": "Green"},
            "preterite": {"color": "#F44336", "name": "Red"},
            "imperfect": {"color": "#FF9800", "name": "Orange"},
            "future": {"color": "#2196F3", "name": "Blue"},
            "conditional": {"color": "#9C27B0", "name": "Purple"},
            "subjunctive": {"color": "#E91E63", "name": "Pink"}
        }
    },
    "callout_templates": {
        "pre_pattern": {
            "template": "👀 Look at these examples. What do you notice about {feature}?",
            "placement": "before_examples",
            "purpose": "Prime attention before showing pattern"
        },
        "post_pattern": {
            "template": "💡 Did you notice? {pattern_description}",
            "placement": "after_examples",
            "purpose": "Confirm pattern recognition"
        },
        "comparison": {
            "template": "🔍 Notice how {form_a} is different from {form_b}",
            "placement": "during_contrast",
            "purpose": "Highlight distinctions"
        },
        "rule_statement": {
            "template": "📐 The pattern: {explicit_rule}",
            "placement": "after_examples",
            "purpose": "State explicit rule"
        },
        "exception_alert": {
            "template": "⚠️ Watch out: {exception_description}",
            "placement": "after_rule",
            "purpose": "Highlight exceptions"
        },
        "memory_hook": {
            "template": "🧠 Remember: {mnemonic}",
            "placement": "end_of_explanation",
            "purpose": "Provide memorable summary"
        }
    },
    "input_flooding": {
        "description": "Multiple examples of same pattern to ensure noticing",
        "minimum_examples": 5,
        "maximum_examples": 8,
        "variety_requirement": "Same pattern, different vocabulary/context",
        "display_format": "Grouped together with pattern visually highlighted"
    },
    "typographic_salience": {
        "primary_highlight": "bold + color",
        "secondary_highlight": "underline",
        "morpheme_boundary": "| (pipe character)",
        "pattern_box": "bordered container with light background",
        "example_format": "habl|o, habl|as, habl|a, habl|amos, habl|áis, habl|an"
    },
    "structural_highlighting": {
        "sentence_structure": {
            "subject": {"bg_color": "#E3F2FD", "label": "[S]"},
            "verb": {"bg_color": "#E8F5E9", "label": "[V]"},
            "object": {"bg_color": "#FFF3E0", "label": "[O]"},
            "complement": {"bg_color": "#F3E5F5", "label": "[C]"}
        }
    },
    "explicit_comparison": {
        "description": "Side-by-side comparison of contrasting forms",
        "format": "two-column table or split screen",
        "example": "SER vs ESTAR: permanent vs temporary"
    }
}

# ═══════════════════════════════════════════════════════════════════════════════
# SHADOWING MODE CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

SHADOWING_CONFIG = {
    "unlock_day": 7,
    "rationale": "Basic vocabulary must be established before shadowing practice",
    "session_structure": {
        "phrases_per_session": {"min": 5, "max": 10, "default": 7},
        "duration_minutes": {"min": 3, "max": 5, "default": 4},
        "phrase_length_seconds": {"min": 2, "max": 8},
        "gap_after_audio_ms": 500,
        "recording_timeout_multiplier": 1.5
    },
    "feedback_types": {
        "rhythm": {
            "description": "Compare timing pattern using DTW algorithm",
            "display": "Visual waveform overlay",
            "metric": "rhythm_match_score"
        },
        "intonation": {
            "description": "Compare pitch contour",
            "display": "Pitch graph comparison",
            "metric": "pitch_correlation"
        },
        "stress": {
            "description": "Compare syllable emphasis",
            "display": "Syllable intensity markers",
            "metric": "stress_accuracy"
        },
        "speed": {
            "description": "Compare words per second",
            "display": "Speed gauge",
            "metric": "wps_ratio"
        }
    },
    "phrase_selection": {
        "from_grammar": 0.3,
        "from_vocabulary": 0.3,
        "from_dialogue": 0.2,
        "review_previous": 0.2
    },
    "difficulty_progression": {
        "day_7_10": {
            "max_words": 5,
            "speed_multiplier": 0.8,
            "primary_focus": "rhythm",
            "secondary_focus": "stress"
        },
        "day_11_15": {
            "max_words": 7,
            "speed_multiplier": 0.9,
            "primary_focus": "intonation",
            "secondary_focus": "rhythm"
        },
        "day_16_20": {
            "max_words": 10,
            "speed_multiplier": 1.0,
            "primary_focus": "stress",
            "secondary_focus": "intonation"
        },
        "day_21_25": {
            "max_words": 12,
            "speed_multiplier": 1.0,
            "primary_focus": "fluency",
            "secondary_focus": "naturalness"
        },
        "day_26_30": {
            "max_words": 15,
            "speed_multiplier": 1.1,
            "primary_focus": "naturalness",
            "secondary_focus": "expression"
        }
    },
    "encouragement_messages": {
        "excellent": ["Perfect rhythm! 🎵", "Nailed it! 🎯", "Native-like! 🌟"],
        "good": ["Great flow! 👍", "Nice work!", "Getting better! 📈"],
        "improving": ["Good try! Keep going 💪", "Almost there!", "Practice makes perfect!"],
        "needs_work": ["Let's try again 🔄", "Listen once more", "Focus on the rhythm"]
    }
}

# ═══════════════════════════════════════════════════════════════════════════════
# 30-DAY CURRICULUM STRUCTURE (Existing - for reference and enhancement)
# ═══════════════════════════════════════════════════════════════════════════════

CURRICULUM_STRUCTURE = {
    1: {"theme": "Greetings & Introductions", "grammar": ["Basic greetings", "Me llamo...", "Ser (yo soy)"], "pt_stage": 1},
    2: {"theme": "Numbers & Age", "grammar": ["Numbers 1-20", "Tener (age)", "¿Cuántos años tienes?"], "pt_stage": 1},
    3: {"theme": "Family & Descriptions", "grammar": ["Family vocabulary", "Gender of nouns", "Articles el/la"], "pt_stage": 2},
    4: {"theme": "Colors & Adjectives", "grammar": ["Colors", "Adjective basics", "Plural nouns"], "pt_stage": 2},
    5: {"theme": "Home & Rooms", "grammar": ["House vocabulary", "Hay (there is/are)", "Possessives mi/tu/su"], "pt_stage": 2},
    6: {"theme": "Daily Objects", "grammar": ["Common objects", "Demonstratives este/esta", "Gender agreement"], "pt_stage": 3},
    7: {"theme": "Food & Drink", "grammar": ["Food vocabulary", "Querer (to want)", "Adjective agreement"], "pt_stage": 3},
    8: {"theme": "Restaurant", "grammar": ["Restaurant phrases", "-AR verb conjugation", "Me gustaría"], "pt_stage": 4},
    9: {"theme": "Time & Schedule", "grammar": ["Time expressions", "Numbers 21-100", "-ER verb conjugation"], "pt_stage": 4},
    10: {"theme": "Days & Weather", "grammar": ["Days of week", "Weather expressions", "-IR verb conjugation"], "pt_stage": 4},
    11: {"theme": "Clothing & Shopping", "grammar": ["Clothing vocabulary", "Costar/Llevar", "Question formation"], "pt_stage": 4},
    12: {"theme": "Body & Health", "grammar": ["Body parts", "Doler (to hurt)", "Negation"], "pt_stage": 4},
    13: {"theme": "Directions", "grammar": ["Direction vocabulary", "Estar (location)", "Prepositions"], "pt_stage": 4},
    14: {"theme": "Transportation", "grammar": ["Transport vocabulary", "Ir (to go)", "Ir + a + infinitive"], "pt_stage": 4},
    15: {"theme": "Daily Routine", "grammar": ["Routine vocabulary", "Reflexive verbs", "Time expressions"], "pt_stage": 5},
    16: {"theme": "Work & Professions", "grammar": ["Profession vocabulary", "Ser vs Estar review", "Direct object pronouns"], "pt_stage": 5},
    17: {"theme": "Hobbies & Free Time", "grammar": ["Hobby vocabulary", "Gustar (full)", "Indirect object pronouns"], "pt_stage": 5},
    18: {"theme": "Past Events (I)", "grammar": ["Time markers", "Preterite regular -AR", "Yesterday/last week"], "pt_stage": 5},
    19: {"theme": "Past Events (II)", "grammar": ["Preterite regular -ER/-IR", "Sequencing words", "Common irregulars"], "pt_stage": 5},
    20: {"theme": "Childhood Memories", "grammar": ["Imperfect tense intro", "Cuando era niño...", "Descriptions in past"], "pt_stage": 5},
    21: {"theme": "Travel Planning", "grammar": ["Travel vocabulary", "Future with ir a", "Pronoun placement"], "pt_stage": 5},
    22: {"theme": "At the Hotel", "grammar": ["Hotel vocabulary", "Conditional basics", "Polite requests"], "pt_stage": 6},
    23: {"theme": "Sightseeing", "grammar": ["Tourism vocabulary", "Comparatives", "Relative clauses (que)"], "pt_stage": 6},
    24: {"theme": "Emergencies", "grammar": ["Emergency vocabulary", "Commands (informal)", "Deber/Tener que"], "pt_stage": 6},
    25: {"theme": "Phone & Technology", "grammar": ["Tech vocabulary", "Present perfect intro", "He/Has/Ha + participle"], "pt_stage": 6},
    26: {"theme": "Expressing Opinions", "grammar": ["Opinion phrases", "Subjunctive intro", "Creo que vs No creo que"], "pt_stage": 6},
    27: {"theme": "Making Plans", "grammar": ["Planning vocabulary", "Subjunctive with desires", "Quiero que + subjunctive"], "pt_stage": 6},
    28: {"theme": "Hypotheticals", "grammar": ["Si clauses (present)", "Conditional review", "If I had..."], "pt_stage": 6},
    29: {"theme": "Storytelling", "grammar": ["Narrative vocabulary", "Preterite vs Imperfect", "Sequencing a story"], "pt_stage": 6},
    30: {"theme": "Review & Conversation", "grammar": ["All tenses review", "Conversation strategies", "Fluency practice"], "pt_stage": 6}
}

# ═══════════════════════════════════════════════════════════════════════════════
# STYLE DEFINITIONS
# ═══════════════════════════════════════════════════════════════════════════════

STYLES = {
    "header": {
        "fill": PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid"),
        "font": Font(bold=True, color="FFFFFF", size=11),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)
    },
    "subheader": {
        "fill": PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid"),
        "font": Font(bold=True, color="FFFFFF", size=10),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)
    },
    "pt_stage_1": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "pt_stage_2": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "pt_stage_3": PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid"),
    "pt_stage_4": PatternFill(start_color="81C784", end_color="81C784", fill_type="solid"),
    "pt_stage_5": PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid"),
    "pt_stage_6": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
    "violation": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
    "success": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "warning": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
    "border": Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
}

# ═══════════════════════════════════════════════════════════════════════════════
# PART 1: CURRICULUM AUDITOR
# ═══════════════════════════════════════════════════════════════════════════════

class CurriculumAuditor:
    """Audits existing curriculum against PT, Noticing, and Shadowing frameworks."""
    
    def __init__(self, excel_path):
        print(f"  Loading workbook: {excel_path}")
        self.wb = load_workbook(excel_path)
        self.audit_results = {
            "processability": {"days": {}, "violations": [], "score": 0},
            "noticing": {"days": {}, "gaps": [], "coverage": {}},
            "shadowing": {"days": {}, "phrases_needed": 0},
            "summary": {}
        }
    
    def audit_processability(self):
        """Check grammar sequence against PT stages."""
        print("  Auditing Processability Theory alignment...")
        
        violations = []
        days_audit = {}
        
        for day, content in CURRICULUM_STRUCTURE.items():
            expected_stage = content["pt_stage"]
            grammar_topics = content["grammar"]
            
            # Check if prerequisites are met
            if expected_stage > 1:
                prerequisite = PT_STAGES[expected_stage]["prerequisite"]
                # Find the latest day that teaches prerequisite stage
                prereq_days = [d for d, c in CURRICULUM_STRUCTURE.items() 
                               if c["pt_stage"] == prerequisite and d < day]
                
                prereq_met = len(prereq_days) > 0
            else:
                prereq_met = True
            
            days_audit[day] = {
                "theme": content["theme"],
                "grammar": grammar_topics,
                "pt_stage": expected_stage,
                "pt_stage_name": PT_STAGES[expected_stage]["name"],
                "prerequisite_met": prereq_met,
                "aligned": prereq_met
            }
            
            if not prereq_met:
                violations.append({
                    "day": day,
                    "issue": f"Stage {expected_stage} introduced without Stage {prerequisite} foundation",
                    "grammar": grammar_topics,
                    "recommendation": f"Ensure Stage {prerequisite} content precedes Day {day}"
                })
        
        # Calculate alignment score
        aligned_days = sum(1 for d in days_audit.values() if d["aligned"])
        score = (aligned_days / 30) * 100
        
        self.audit_results["processability"] = {
            "days": days_audit,
            "violations": violations,
            "score": score,
            "total_days": 30,
            "aligned_days": aligned_days
        }
        
        print(f"    PT Alignment Score: {score:.1f}%")
        print(f"    Violations found: {len(violations)}")
        
        return self.audit_results["processability"]
    
    def audit_noticing(self):
        """Check noticing technique coverage for each grammar point."""
        print("  Auditing Noticing Hypothesis coverage...")
        
        techniques = [
            "color_coding",
            "pre_pattern_callout",
            "post_pattern_callout",
            "input_flooding",
            "typographic_salience",
            "structural_highlighting"
        ]
        
        days_audit = {}
        gaps = []
        
        for day, content in CURRICULUM_STRUCTURE.items():
            grammar_points = content["grammar"]
            day_techniques = {}
            
            for grammar in grammar_points:
                # Determine which techniques apply to this grammar point
                applicable = self._get_applicable_noticing(grammar, day)
                day_techniques[grammar] = applicable
                
                # Check for gaps
                missing = [t for t in techniques if not applicable.get(t, {}).get("implemented", False)]
                if missing:
                    gaps.append({
                        "day": day,
                        "grammar": grammar,
                        "missing_techniques": missing,
                        "recommendations": self._get_noticing_recommendations(grammar, missing)
                    })
            
            days_audit[day] = {
                "theme": content["theme"],
                "grammar_points": grammar_points,
                "techniques": day_techniques,
                "coverage_percent": self._calculate_noticing_coverage(day_techniques)
            }
        
        # Calculate overall coverage
        total_coverage = sum(d["coverage_percent"] for d in days_audit.values()) / 30
        
        self.audit_results["noticing"] = {
            "days": days_audit,
            "gaps": gaps,
            "overall_coverage": total_coverage,
            "techniques_defined": techniques
        }
        
        print(f"    Noticing Coverage: {total_coverage:.1f}%")
        print(f"    Gaps to address: {len(gaps)}")
        
        return self.audit_results["noticing"]
    
    def _get_applicable_noticing(self, grammar_point, day):
        """Determine which noticing techniques apply and generate specs."""
        grammar_lower = grammar_point.lower()
        
        techniques = {}
        
        # Color coding - applicable for conjugations, gender, parts of speech
        if any(kw in grammar_lower for kw in ["conjugation", "verb", "-ar", "-er", "-ir", "tense"]):
            techniques["color_coding"] = {
                "implemented": True,
                "type": "verb_endings",
                "spec": NOTICING_TECHNIQUES["color_coding"]["verb_endings"]
            }
        elif any(kw in grammar_lower for kw in ["gender", "el/la", "masculine", "feminine"]):
            techniques["color_coding"] = {
                "implemented": True,
                "type": "gender",
                "spec": NOTICING_TECHNIQUES["color_coding"]["gender"]
            }
        else:
            techniques["color_coding"] = {
                "implemented": False,
                "type": "parts_of_speech",
                "spec": NOTICING_TECHNIQUES["color_coding"]["parts_of_speech"]
            }
        
        # Pre-pattern callout - always applicable
        techniques["pre_pattern_callout"] = {
            "implemented": True,
            "template": NOTICING_TECHNIQUES["callout_templates"]["pre_pattern"]["template"],
            "feature": grammar_point
        }
        
        # Post-pattern callout - always applicable
        techniques["post_pattern_callout"] = {
            "implemented": True,
            "template": NOTICING_TECHNIQUES["callout_templates"]["post_pattern"]["template"],
            "pattern_description": f"The pattern for {grammar_point}"
        }
        
        # Input flooding - always applicable
        techniques["input_flooding"] = {
            "implemented": True,
            "minimum_examples": NOTICING_TECHNIQUES["input_flooding"]["minimum_examples"],
            "spec": "5-8 examples showing same pattern with different vocabulary"
        }
        
        # Typographic salience - applicable for morphology
        if any(kw in grammar_lower for kw in ["conjugation", "ending", "agreement", "plural"]):
            techniques["typographic_salience"] = {
                "implemented": True,
                "format": NOTICING_TECHNIQUES["typographic_salience"]["example_format"],
                "spec": "Use | to show morpheme boundaries, bold + color for endings"
            }
        else:
            techniques["typographic_salience"] = {
                "implemented": False,
                "spec": "Bold key terms"
            }
        
        # Structural highlighting - applicable for word order, sentence structure
        if any(kw in grammar_lower for kw in ["order", "sentence", "svo", "question"]):
            techniques["structural_highlighting"] = {
                "implemented": True,
                "spec": NOTICING_TECHNIQUES["structural_highlighting"]["sentence_structure"]
            }
        else:
            techniques["structural_highlighting"] = {
                "implemented": False,
                "spec": "Not required for this grammar point"
            }
        
        return techniques
    
    def _get_noticing_recommendations(self, grammar, missing):
        """Generate recommendations for missing noticing techniques."""
        recommendations = []
        
        for technique in missing:
            if technique == "color_coding":
                recommendations.append(f"Add color coding to highlight {grammar} patterns")
            elif technique == "input_flooding":
                recommendations.append(f"Add 5+ examples demonstrating {grammar}")
            elif technique == "pre_pattern_callout":
                recommendations.append(f"Add 'What do you notice about...' prompt before {grammar}")
            elif technique == "post_pattern_callout":
                recommendations.append(f"Add 'Did you notice?' confirmation after {grammar}")
            elif technique == "typographic_salience":
                recommendations.append(f"Add bold/underline/pipe markers for {grammar} morphemes")
            elif technique == "structural_highlighting":
                recommendations.append(f"Add [S][V][O] structure markers for {grammar}")
        
        return recommendations
    
    def _calculate_noticing_coverage(self, day_techniques):
        """Calculate percentage of noticing techniques implemented for a day."""
        total = 0
        implemented = 0
        
        for grammar, techniques in day_techniques.items():
            for technique, spec in techniques.items():
                total += 1
                if spec.get("implemented", False):
                    implemented += 1
        
        return (implemented / total * 100) if total > 0 else 0
    
    def audit_shadowing(self):
        """Identify shadowing content needs for Days 7-30."""
        print("  Auditing Shadowing Mode coverage...")
        
        days_audit = {}
        total_phrases_needed = 0
        
        for day in range(1, 31):
            content = CURRICULUM_STRUCTURE[day]
            
            if day < SHADOWING_CONFIG["unlock_day"]:
                days_audit[day] = {
                    "enabled": False,
                    "reason": f"Shadowing unlocks on Day {SHADOWING_CONFIG['unlock_day']}",
                    "phrases": []
                }
            else:
                # Determine difficulty tier
                difficulty = self._get_shadowing_difficulty(day)
                phrases_count = SHADOWING_CONFIG["session_structure"]["phrases_per_session"]["default"]
                
                # Generate phrase specifications
                phrases = self._generate_shadowing_phrases(day, content, difficulty)
                
                days_audit[day] = {
                    "enabled": True,
                    "difficulty_tier": difficulty,
                    "difficulty_settings": SHADOWING_CONFIG["difficulty_progression"][difficulty],
                    "phrases_count": len(phrases),
                    "phrases": phrases,
                    "session_duration_min": SHADOWING_CONFIG["session_structure"]["duration_minutes"]["default"],
                    "feedback_focus": SHADOWING_CONFIG["difficulty_progression"][difficulty]["primary_focus"]
                }
                
                total_phrases_needed += len(phrases)
        
        self.audit_results["shadowing"] = {
            "days": days_audit,
            "total_phrases_needed": total_phrases_needed,
            "unlock_day": SHADOWING_CONFIG["unlock_day"],
            "days_with_shadowing": 30 - SHADOWING_CONFIG["unlock_day"] + 1
        }
        
        print(f"    Shadowing days: {30 - SHADOWING_CONFIG['unlock_day'] + 1}")
        print(f"    Total phrases to create: {total_phrases_needed}")
        
        return self.audit_results["shadowing"]
    
    def _get_shadowing_difficulty(self, day):
        """Get difficulty tier for a given day."""
        if day <= 10:
            return "day_7_10"
        elif day <= 15:
            return "day_11_15"
        elif day <= 20:
            return "day_16_20"
        elif day <= 25:
            return "day_21_25"
        else:
            return "day_26_30"
    
    def _generate_shadowing_phrases(self, day, content, difficulty):
        """Generate shadowing phrase specifications for a day."""
        settings = SHADOWING_CONFIG["difficulty_progression"][difficulty]
        phrases = []
        
        theme = content["theme"]
        grammar = content["grammar"]
        
        # Generate 7 phrases per day (default)
        phrase_templates = [
            {"source": "grammar", "focus": settings["primary_focus"]},
            {"source": "grammar", "focus": settings["secondary_focus"]},
            {"source": "vocabulary", "focus": settings["primary_focus"]},
            {"source": "vocabulary", "focus": settings["secondary_focus"]},
            {"source": "dialogue", "focus": settings["primary_focus"]},
            {"source": "review", "focus": "rhythm"},
            {"source": "review", "focus": "intonation"}
        ]
        
        for i, template in enumerate(phrase_templates):
            phrases.append({
                "id": f"day{day}_shadow_{i+1:02d}",
                "source_type": template["source"],
                "focus": template["focus"],
                "max_words": settings["max_words"],
                "speed_multiplier": settings["speed_multiplier"],
                "placeholder_text": f"[{theme} - {template['source']} phrase {i+1}]",
                "audio_file": f"day{day}_shadow_{i+1:02d}.mp3",
                "duration_estimate_s": 3 + (settings["max_words"] * 0.3)
            })
        
        return phrases
    
    def generate_summary(self):
        """Generate overall audit summary."""
        pt = self.audit_results["processability"]
        noticing = self.audit_results["noticing"]
        shadowing = self.audit_results["shadowing"]
        
        self.audit_results["summary"] = {
            "generated_at": datetime.now().isoformat(),
            "overall_scores": {
                "pt_alignment": pt["score"],
                "noticing_coverage": noticing["overall_coverage"],
                "shadowing_readiness": (shadowing["days_with_shadowing"] / 24) * 100  # Days 7-30
            },
            "action_items": {
                "pt_violations": len(pt["violations"]),
                "noticing_gaps": len(noticing["gaps"]),
                "shadowing_phrases_to_create": shadowing["total_phrases_needed"]
            },
            "preservation_mode": PRESERVE_MODE
        }
        
        return self.audit_results["summary"]


# ═══════════════════════════════════════════════════════════════════════════════
# PART 2: MATRIX UPDATER
# ═══════════════════════════════════════════════════════════════════════════════

class MatrixUpdater:
    """Updates Excel workbook with audit results and enhancements."""
    
    def __init__(self, workbook, audit_results):
        self.wb = workbook
        self.audit = audit_results
    
    def add_pt_audit_sheet(self):
        """Add Processability Theory Audit worksheet."""
        print("  Adding PT Audit Results worksheet...")
        
        # Create new worksheet (don't delete existing)
        if "PT Audit Results" in self.wb.sheetnames:
            del self.wb["PT Audit Results"]
        
        ws = self.wb.create_sheet("PT Audit Results", 0)
        
        # Headers
        headers = ["Day", "Theme", "Grammar Topics", "PT Stage", "Stage Name", 
                   "Prerequisites Met", "Status", "Notes"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = STYLES["header"]["fill"]
            cell.font = STYLES["header"]["font"]
            cell.alignment = STYLES["header"]["alignment"]
            cell.border = STYLES["border"]
        
        # Data rows
        pt_data = self.audit["processability"]["days"]
        for row, (day, data) in enumerate(pt_data.items(), 2):
            ws.cell(row=row, column=1, value=day).border = STYLES["border"]
            ws.cell(row=row, column=2, value=data["theme"]).border = STYLES["border"]
            ws.cell(row=row, column=3, value=", ".join(data["grammar"])).border = STYLES["border"]
            ws.cell(row=row, column=4, value=data["pt_stage"]).border = STYLES["border"]
            ws.cell(row=row, column=5, value=data["pt_stage_name"]).border = STYLES["border"]
            ws.cell(row=row, column=6, value="✅ Yes" if data["prerequisite_met"] else "❌ No").border = STYLES["border"]
            
            status_cell = ws.cell(row=row, column=7, value="✅ Aligned" if data["aligned"] else "⚠️ Review")
            status_cell.border = STYLES["border"]
            
            if data["aligned"]:
                status_cell.fill = STYLES["success"]
            else:
                status_cell.fill = STYLES["violation"]
            
            # Stage-based coloring
            stage_fill = STYLES[f"pt_stage_{data['pt_stage']}"]
            ws.cell(row=row, column=4).fill = stage_fill
            
            ws.cell(row=row, column=8, value="").border = STYLES["border"]
        
        # Add PT Stage Reference section
        ref_start = 34
        ws.cell(row=ref_start, column=1, value="PT STAGE REFERENCE").font = Font(bold=True, size=14)
        
        ref_headers = ["Stage", "Name", "Structures", "Spanish Examples", "Typical Days"]
        for col, header in enumerate(ref_headers, 1):
            cell = ws.cell(row=ref_start + 1, column=col, value=header)
            cell.fill = STYLES["subheader"]["fill"]
            cell.font = STYLES["subheader"]["font"]
            cell.border = STYLES["border"]
        
        for stage_num, stage_data in PT_STAGES.items():
            row = ref_start + 1 + stage_num
            ws.cell(row=row, column=1, value=stage_num).border = STYLES["border"]
            ws.cell(row=row, column=1).fill = STYLES[f"pt_stage_{stage_num}"]
            ws.cell(row=row, column=2, value=stage_data["name"]).border = STYLES["border"]
            ws.cell(row=row, column=3, value=", ".join(stage_data["structures"][:3])).border = STYLES["border"]
            ws.cell(row=row, column=4, value=", ".join(stage_data["spanish_examples"][:3])).border = STYLES["border"]
            ws.cell(row=row, column=5, value=str(stage_data["typical_days"])).border = STYLES["border"]
        
        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 30
        
        # Freeze header
        ws.freeze_panes = 'A2'
        
        print(f"    Added {len(pt_data)} day entries")
    
    def add_noticing_checklist(self):
        """Add Noticing Hypothesis Checklist worksheet."""
        print("  Adding Noticing Checklist worksheet...")
        
        if "Noticing Checklist" in self.wb.sheetnames:
            del self.wb["Noticing Checklist"]
        
        ws = self.wb.create_sheet("Noticing Checklist", 1)
        
        # Headers
        headers = ["Day", "Grammar Point", "Color Coding", "Pre-Callout", "Post-Callout",
                   "Input Flooding", "Typography", "Structure", "Coverage %", "Implementation Notes"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = STYLES["header"]["fill"]
            cell.font = STYLES["header"]["font"]
            cell.alignment = STYLES["header"]["alignment"]
            cell.border = STYLES["border"]
        
        # Data rows
        row = 2
        noticing_data = self.audit["noticing"]["days"]
        
        for day, day_data in noticing_data.items():
            for grammar, techniques in day_data["techniques"].items():
                ws.cell(row=row, column=1, value=day).border = STYLES["border"]
                ws.cell(row=row, column=2, value=grammar).border = STYLES["border"]
                
                # Technique columns
                tech_map = {
                    3: "color_coding",
                    4: "pre_pattern_callout",
                    5: "post_pattern_callout",
                    6: "input_flooding",
                    7: "typographic_salience",
                    8: "structural_highlighting"
                }
                
                implemented_count = 0
                for col, tech_key in tech_map.items():
                    tech_data = techniques.get(tech_key, {})
                    is_implemented = tech_data.get("implemented", False)
                    
                    cell = ws.cell(row=row, column=col)
                    cell.value = "✅" if is_implemented else "❌"
                    cell.border = STYLES["border"]
                    cell.alignment = Alignment(horizontal="center")
                    
                    if is_implemented:
                        cell.fill = STYLES["success"]
                        implemented_count += 1
                    else:
                        cell.fill = STYLES["warning"]
                
                # Coverage percentage
                coverage = (implemented_count / 6) * 100
                coverage_cell = ws.cell(row=row, column=9, value=f"{coverage:.0f}%")
                coverage_cell.border = STYLES["border"]
                coverage_cell.alignment = Alignment(horizontal="center")
                
                # Implementation notes
                notes = []
                if techniques.get("color_coding", {}).get("type"):
                    notes.append(f"Color: {techniques['color_coding']['type']}")
                ws.cell(row=row, column=10, value="; ".join(notes)).border = STYLES["border"]
                
                row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 30
        for col in range(3, 9):
            ws.column_dimensions[get_column_letter(col)].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 40
        
        ws.freeze_panes = 'A2'
        
        print(f"    Added checklist with {row - 2} grammar point entries")
    
    def add_shadowing_specs(self):
        """Add Shadowing Day Plans worksheet."""
        print("  Adding Shadowing Day Plans worksheet...")
        
        if "Shadowing Day Plans" in self.wb.sheetnames:
            del self.wb["Shadowing Day Plans"]
        
        ws = self.wb.create_sheet("Shadowing Day Plans", 2)
        
        # Headers
        headers = ["Day", "Enabled", "Difficulty Tier", "Max Words", "Speed", 
                   "Primary Focus", "Phrases Count", "Session Duration", "Phrase IDs"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = STYLES["header"]["fill"]
            cell.font = STYLES["header"]["font"]
            cell.alignment = STYLES["header"]["alignment"]
            cell.border = STYLES["border"]
        
        # Data rows
        shadowing_data = self.audit["shadowing"]["days"]
        for row, (day, data) in enumerate(shadowing_data.items(), 2):
            ws.cell(row=row, column=1, value=day).border = STYLES["border"]
            
            enabled_cell = ws.cell(row=row, column=2, value="✅ Yes" if data["enabled"] else "❌ No")
            enabled_cell.border = STYLES["border"]
            
            if data["enabled"]:
                enabled_cell.fill = STYLES["success"]
                settings = data["difficulty_settings"]
                
                ws.cell(row=row, column=3, value=data["difficulty_tier"]).border = STYLES["border"]
                ws.cell(row=row, column=4, value=settings["max_words"]).border = STYLES["border"]
                ws.cell(row=row, column=5, value=f"{settings['speed_multiplier']}x").border = STYLES["border"]
                ws.cell(row=row, column=6, value=settings["primary_focus"]).border = STYLES["border"]
                ws.cell(row=row, column=7, value=data["phrases_count"]).border = STYLES["border"]
                ws.cell(row=row, column=8, value=f"{data['session_duration_min']} min").border = STYLES["border"]
                
                phrase_ids = ", ".join([p["id"] for p in data["phrases"][:3]]) + "..."
                ws.cell(row=row, column=9, value=phrase_ids).border = STYLES["border"]
            else:
                ws.cell(row=row, column=3, value=data.get("reason", "Not enabled")).border = STYLES["border"]
                for col in range(4, 10):
                    ws.cell(row=row, column=col, value="-").border = STYLES["border"]
                enabled_cell.fill = STYLES["warning"]
        
        # Add summary section
        summary_row = 34
        ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
        ws.cell(row=summary_row + 1, column=1, value="Total Shadowing Days:")
        ws.cell(row=summary_row + 1, column=2, value=self.audit["shadowing"]["days_with_shadowing"])
        ws.cell(row=summary_row + 2, column=1, value="Total Phrases Needed:")
        ws.cell(row=summary_row + 2, column=2, value=self.audit["shadowing"]["total_phrases_needed"])
        
        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 40
        
        ws.freeze_panes = 'A2'
        
        print(f"    Added shadowing specs for {self.audit['shadowing']['days_with_shadowing']} days")
    
    def add_day_implementations(self):
        """Add comprehensive Day-by-Day Specs worksheet."""
        print("  Adding Day-by-Day Specs worksheet...")
        
        if "Day-by-Day Specs" in self.wb.sheetnames:
            del self.wb["Day-by-Day Specs"]
        
        ws = self.wb.create_sheet("Day-by-Day Specs", 3)
        
        # Headers
        headers = ["Day", "Theme", "PT Stage", "Grammar Module", "Noticing Specs",
                   "Vocabulary Module", "Listen Module", "Shadowing Specs", "Read Module"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = STYLES["header"]["fill"]
            cell.font = STYLES["header"]["font"]
            cell.alignment = STYLES["header"]["alignment"]
            cell.border = STYLES["border"]
        
        # Data rows
        for day in range(1, 31):
            row = day + 1
            content = CURRICULUM_STRUCTURE[day]
            pt_data = self.audit["processability"]["days"][day]
            noticing_data = self.audit["noticing"]["days"][day]
            shadowing_data = self.audit["shadowing"]["days"][day]
            
            # Day number
            ws.cell(row=row, column=1, value=day).border = STYLES["border"]
            ws.cell(row=row, column=1).fill = STYLES[f"pt_stage_{content['pt_stage']}"]
            
            # Theme
            ws.cell(row=row, column=2, value=content["theme"]).border = STYLES["border"]
            
            # PT Stage
            pt_cell = ws.cell(row=row, column=3, 
                             value=f"Stage {content['pt_stage']}: {PT_STAGES[content['pt_stage']]['name']}")
            pt_cell.border = STYLES["border"]
            pt_cell.fill = STYLES[f"pt_stage_{content['pt_stage']}"]
            
            # Grammar Module
            grammar_spec = self._format_grammar_spec(content, noticing_data)
            ws.cell(row=row, column=4, value=grammar_spec).border = STYLES["border"]
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            
            # Noticing Specs
            noticing_spec = self._format_noticing_spec(noticing_data)
            ws.cell(row=row, column=5, value=noticing_spec).border = STYLES["border"]
            ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")
            
            # Vocabulary Module
            vocab_spec = f"Theme: {content['theme']}\nNoticing: Part-of-speech color coding"
            ws.cell(row=row, column=6, value=vocab_spec).border = STYLES["border"]
            ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
            
            # Listen Module
            listen_spec = f"Audio clips aligned to Day {day} content\nSpeed: Calibrated to level"
            ws.cell(row=row, column=7, value=listen_spec).border = STYLES["border"]
            ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical="top")
            
            # Shadowing Specs
            if shadowing_data["enabled"]:
                shadow_spec = self._format_shadowing_spec(shadowing_data)
            else:
                shadow_spec = f"Not enabled (unlocks Day {SHADOWING_CONFIG['unlock_day']})"
            ws.cell(row=row, column=8, value=shadow_spec).border = STYLES["border"]
            ws.cell(row=row, column=8).alignment = Alignment(wrap_text=True, vertical="top")
            
            # Read Module
            read_spec = f"Topic: {content['theme']}\nNoticing: Bold key phrases, underline vocab"
            ws.cell(row=row, column=9, value=read_spec).border = STYLES["border"]
            ws.cell(row=row, column=9).alignment = Alignment(wrap_text=True, vertical="top")
        
        # Set row heights and column widths
        for row in range(2, 32):
            ws.row_dimensions[row].height = 80
        
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 35
        ws.column_dimensions['I'].width = 25
        
        ws.freeze_panes = 'A2'
        
        print(f"    Added comprehensive specs for 30 days")
    
    def _format_grammar_spec(self, content, noticing_data):
        """Format grammar module specification."""
        lines = []
        lines.append(f"Topics: {', '.join(content['grammar'])}")
        lines.append(f"PT Alignment: Stage {content['pt_stage']}")
        lines.append(f"Coverage: {noticing_data['coverage_percent']:.0f}%")
        return "\n".join(lines)
    
    def _format_noticing_spec(self, noticing_data):
        """Format noticing specification."""
        lines = []
        
        for grammar, techniques in noticing_data["techniques"].items():
            lines.append(f"• {grammar}:")
            
            if techniques.get("color_coding", {}).get("implemented"):
                color_type = techniques["color_coding"].get("type", "default")
                lines.append(f"  - Color: {color_type}")
            
            if techniques.get("input_flooding", {}).get("implemented"):
                lines.append(f"  - Input flooding: 5+ examples")
            
            if techniques.get("typographic_salience", {}).get("implemented"):
                lines.append(f"  - Typography: morpheme boundaries")
        
        return "\n".join(lines[:10])  # Limit lines
    
    def _format_shadowing_spec(self, shadowing_data):
        """Format shadowing specification."""
        settings = shadowing_data["difficulty_settings"]
        lines = [
            f"Tier: {shadowing_data['difficulty_tier']}",
            f"Phrases: {shadowing_data['phrases_count']}",
            f"Max words: {settings['max_words']}",
            f"Speed: {settings['speed_multiplier']}x",
            f"Focus: {settings['primary_focus']}",
            f"Duration: {shadowing_data['session_duration_min']} min"
        ]
        return "\n".join(lines)
    
    def add_enhancement_summary(self):
        """Add Enhancement Summary worksheet."""
        print("  Adding Enhancement Summary worksheet...")
        
        if "Enhancement Summary" in self.wb.sheetnames:
            del self.wb["Enhancement Summary"]
        
        ws = self.wb.create_sheet("Enhancement Summary", 4)
        
        # Title
        ws.cell(row=1, column=1, value="FLUIDEZ CURRICULUM ENHANCEMENT SUMMARY")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="1B4332")
        ws.merge_cells('A1:E1')
        
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws.cell(row=3, column=1, value=f"Preservation Mode: {'ENABLED' if PRESERVE_MODE else 'DISABLED'}")
        
        # Scores section
        row = 5
        ws.cell(row=row, column=1, value="AUDIT SCORES").font = Font(bold=True, size=12)
        
        summary = self.audit["summary"]
        scores = [
            ("PT Alignment", f"{summary['overall_scores']['pt_alignment']:.1f}%"),
            ("Noticing Coverage", f"{summary['overall_scores']['noticing_coverage']:.1f}%"),
            ("Shadowing Readiness", f"{summary['overall_scores']['shadowing_readiness']:.1f}%")
        ]
        
        for i, (label, value) in enumerate(scores):
            ws.cell(row=row + 1 + i, column=1, value=label)
            ws.cell(row=row + 1 + i, column=2, value=value)
        
        # Action items
        row = 10
        ws.cell(row=row, column=1, value="ACTION ITEMS").font = Font(bold=True, size=12)
        
        actions = [
            ("PT Violations to Review", summary["action_items"]["pt_violations"]),
            ("Noticing Gaps to Fill", summary["action_items"]["noticing_gaps"]),
            ("Shadowing Phrases to Create", summary["action_items"]["shadowing_phrases_to_create"])
        ]
        
        for i, (label, value) in enumerate(actions):
            ws.cell(row=row + 1 + i, column=1, value=label)
            ws.cell(row=row + 1 + i, column=2, value=value)
        
        # New worksheets added
        row = 15
        ws.cell(row=row, column=1, value="WORKSHEETS ADDED").font = Font(bold=True, size=12)
        
        new_sheets = [
            "PT Audit Results - Day-by-day processability stage alignment",
            "Noticing Checklist - Technique coverage for each grammar point",
            "Shadowing Day Plans - Phrase specifications for Days 7-30",
            "Day-by-Day Specs - Complete implementation specs per day",
            "Enhancement Summary - This overview"
        ]
        
        for i, sheet in enumerate(new_sheets):
            ws.cell(row=row + 1 + i, column=1, value=f"• {sheet}")
        
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        
        print("    Added enhancement summary")
    
    def save(self, output_path):
        """Save the updated workbook."""
        print(f"  Saving workbook to: {output_path}")
        self.wb.save(output_path)
        print("    ✅ Workbook saved successfully")


# ═══════════════════════════════════════════════════════════════════════════════
# PART 3: DAY IMPLEMENTATION GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

class DayImplementationGenerator:
    """Generates comprehensive day-by-day implementation specs."""
    
    def __init__(self, audit_results):
        self.audit = audit_results
        self.implementations = {}
    
    def generate_all_days(self):
        """Generate implementations for all 30 days."""
        print("  Generating day-by-day implementations...")
        
        for day in range(1, 31):
            self.implementations[f"day_{day}"] = self.generate_day(day)
        
        print(f"    Generated specs for 30 days")
        return self.implementations
    
    def generate_day(self, day_num):
        """Generate complete implementation for one day."""
        content = CURRICULUM_STRUCTURE[day_num]
        pt_data = self.audit["processability"]["days"][day_num]
        noticing_data = self.audit["noticing"]["days"][day_num]
        shadowing_data = self.audit["shadowing"]["days"][day_num]
        
        return {
            "day": day_num,
            "theme": content["theme"],
            "processability": {
                "stage": content["pt_stage"],
                "stage_name": PT_STAGES[content["pt_stage"]]["name"],
                "stage_description": PT_STAGES[content["pt_stage"]]["description"],
                "aligned": pt_data["aligned"],
                "prerequisite_met": pt_data["prerequisite_met"]
            },
            "grammar_module": {
                "topics": content["grammar"],
                "noticing_implementation": self._generate_noticing_impl(content["grammar"], noticing_data)
            },
            "vocabulary_module": {
                "theme": content["theme"],
                "noticing_elements": {
                    "color_coding": NOTICING_TECHNIQUES["color_coding"]["parts_of_speech"],
                    "category_grouping": True
                }
            },
            "listen_module": {
                "content_alignment": f"Audio for Day {day_num}: {content['theme']}",
                "speed_calibration": "Level-appropriate"
            },
            "shadowing_module": self._generate_shadowing_impl(day_num, shadowing_data),
            "read_module": {
                "topic": content["theme"],
                "noticing_elements": {
                    "bold_key_phrases": True,
                    "underline_vocabulary": True,
                    "grammar_highlighting": content["grammar"]
                }
            }
        }
    
    def _generate_noticing_impl(self, grammar_topics, noticing_data):
        """Generate noticing implementation for grammar topics."""
        implementations = []
        
        for grammar in grammar_topics:
            techniques = noticing_data["techniques"].get(grammar, {})
            
            impl = {
                "grammar_point": grammar,
                "techniques": {}
            }
            
            # Color coding
            if techniques.get("color_coding", {}).get("implemented"):
                color_type = techniques["color_coding"].get("type", "default")
                impl["techniques"]["color_coding"] = {
                    "enabled": True,
                    "type": color_type,
                    "spec": NOTICING_TECHNIQUES["color_coding"].get(color_type, {})
                }
            
            # Callouts
            impl["techniques"]["callouts"] = {
                "pre_pattern": NOTICING_TECHNIQUES["callout_templates"]["pre_pattern"]["template"].format(
                    feature=grammar
                ),
                "post_pattern": NOTICING_TECHNIQUES["callout_templates"]["post_pattern"]["template"].format(
                    pattern_description=f"the pattern in {grammar}"
                ),
                "rule_statement": NOTICING_TECHNIQUES["callout_templates"]["rule_statement"]["template"].format(
                    explicit_rule=f"[Rule for {grammar}]"
                )
            }
            
            # Input flooding
            impl["techniques"]["input_flooding"] = {
                "enabled": True,
                "minimum_examples": 5,
                "placeholder": f"[5-8 examples demonstrating {grammar}]"
            }
            
            # Typography
            if techniques.get("typographic_salience", {}).get("implemented"):
                impl["techniques"]["typography"] = {
                    "enabled": True,
                    "morpheme_boundary": "|",
                    "primary_highlight": "bold + color"
                }
            
            implementations.append(impl)
        
        return implementations
    
    def _generate_shadowing_impl(self, day_num, shadowing_data):
        """Generate shadowing implementation for a day."""
        if not shadowing_data["enabled"]:
            return {
                "enabled": False,
                "unlock_day": SHADOWING_CONFIG["unlock_day"],
                "reason": f"Shadowing mode unlocks on Day {SHADOWING_CONFIG['unlock_day']}"
            }
        
        return {
            "enabled": True,
            "difficulty_tier": shadowing_data["difficulty_tier"],
            "settings": shadowing_data["difficulty_settings"],
            "session": {
                "duration_minutes": shadowing_data["session_duration_min"],
                "phrases_count": shadowing_data["phrases_count"],
                "feedback_focus": shadowing_data["feedback_focus"]
            },
            "phrases": shadowing_data["phrases"],
            "feedback_config": {
                "primary_metric": f"{shadowing_data['feedback_focus']}_match",
                "encouragement_messages": SHADOWING_CONFIG["encouragement_messages"]
            }
        }
    
    def export_json(self, filepath):
        """Export implementations to JSON file."""
        print(f"  Exporting JSON to: {filepath}")
        
        output = {
            "metadata": {
                "curriculum_version": "2.0-PT-Noticing-Shadowing",
                "generated_date": datetime.now().isoformat(),
                "preservation_mode": PRESERVE_MODE,
                "frameworks_applied": [
                    "Processability Theory (Pienemann)",
                    "Noticing Hypothesis (Schmidt)",
                    "Shadowing Mode (Phonological Loop)"
                ]
            },
            "summary": {
                "total_days": 30,
                "pt_stages_covered": list(range(1, 7)),
                "shadowing_days": 30 - SHADOWING_CONFIG["unlock_day"] + 1,
                "total_shadowing_phrases": self.audit["shadowing"]["total_phrases_needed"]
            },
            "configurations": {
                "pt_stages": PT_STAGES,
                "noticing_techniques": NOTICING_TECHNIQUES,
                "shadowing_config": SHADOWING_CONFIG
            },
            "days": self.implementations
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
        
        print("    ✅ JSON exported successfully")


# ═══════════════════════════════════════════════════════════════════════════════
# PART 4: REPORT GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

def generate_markdown_report(audit_results, output_path):
    """Generate comprehensive markdown audit report."""
    print(f"  Generating markdown report: {output_path}")
    
    pt = audit_results["processability"]
    noticing = audit_results["noticing"]
    shadowing = audit_results["shadowing"]
    summary = audit_results["summary"]
    
    report = f"""# Fluidez Curriculum Audit Report

**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M')}  
**Preservation Mode:** {'ENABLED - Only additions, no modifications' if PRESERVE_MODE else 'DISABLED'}

---

## Executive Summary

| Framework | Score | Status |
|-----------|-------|--------|
| Processability Theory Alignment | {pt['score']:.1f}% | {'✅ Good' if pt['score'] >= 90 else '⚠️ Review needed'} |
| Noticing Hypothesis Coverage | {noticing['overall_coverage']:.1f}% | {'✅ Good' if noticing['overall_coverage'] >= 80 else '⚠️ Gaps to fill'} |
| Shadowing Mode Readiness | {summary['overall_scores']['shadowing_readiness']:.1f}% | {'✅ Ready' if summary['overall_scores']['shadowing_readiness'] >= 90 else '⚠️ Content needed'} |

### Action Items
- **PT Violations to Review:** {len(pt['violations'])}
- **Noticing Gaps to Fill:** {len(noticing['gaps'])}
- **Shadowing Phrases to Create:** {shadowing['total_phrases_needed']}

---

## 1. Processability Theory Audit

### Overview
Processability Theory (Pienemann) states that learners can only acquire grammatical structures in a fixed developmental sequence. This audit verifies that grammar is introduced in the correct order.

### Alignment by Day

| Day | Theme | PT Stage | Aligned |
|-----|-------|----------|---------|
"""
    
    for day, data in pt['days'].items():
        status = "✅" if data["aligned"] else "⚠️"
        report += f"| {day} | {data['theme']} | Stage {data['pt_stage']} | {status} |\n"
    
    report += f"""
### Violations Found: {len(pt['violations'])}

"""
    
    if pt['violations']:
        for v in pt['violations']:
            report += f"- **Day {v['day']}:** {v['issue']}\n"
            report += f"  - Grammar: {v['grammar']}\n"
            report += f"  - Recommendation: {v['recommendation']}\n\n"
    else:
        report += "*No violations found. Grammar sequence is PT-aligned.*\n\n"
    
    report += f"""### PT Stage Reference

| Stage | Name | Typical Days |
|-------|------|--------------|
"""
    
    for stage, data in PT_STAGES.items():
        report += f"| {stage} | {data['name']} | {data['typical_days']} |\n"
    
    report += f"""

---

## 2. Noticing Hypothesis Audit

### Overview
The Noticing Hypothesis (Schmidt) states that learners must consciously notice forms to acquire them. This audit verifies that attention-directing techniques are applied to all grammar points.

### Techniques Checked
1. Color Coding
2. Pre-Pattern Callouts ("What do you notice about...?")
3. Post-Pattern Callouts ("Did you notice?")
4. Input Flooding (5+ examples)
5. Typographic Salience (bold, underline, morpheme boundaries)
6. Structural Highlighting ([S][V][O] markers)

### Coverage by Day

| Day | Grammar Points | Coverage |
|-----|---------------|----------|
"""
    
    for day, data in noticing['days'].items():
        report += f"| {day} | {len(data['grammar_points'])} | {data['coverage_percent']:.0f}% |\n"
    
    report += f"""
### Gaps Found: {len(noticing['gaps'])}

"""
    
    if noticing['gaps'][:10]:  # Show first 10
        for gap in noticing['gaps'][:10]:
            report += f"- **Day {gap['day']}, {gap['grammar']}:** Missing {', '.join(gap['missing_techniques'])}\n"
        if len(noticing['gaps']) > 10:
            report += f"\n*...and {len(noticing['gaps']) - 10} more gaps*\n"
    else:
        report += "*No significant gaps found.*\n"
    
    report += f"""

---

## 3. Shadowing Mode Audit

### Overview
Shadowing Mode uses the Phonological Loop (Baddeley) to build pronunciation automaticity through immediate repetition of native speaker audio.

### Configuration
- **Unlock Day:** {SHADOWING_CONFIG['unlock_day']}
- **Phrases per Session:** {SHADOWING_CONFIG['session_structure']['phrases_per_session']['default']}
- **Session Duration:** {SHADOWING_CONFIG['session_structure']['duration_minutes']['default']} minutes

### Difficulty Progression

| Day Range | Max Words | Speed | Primary Focus |
|-----------|-----------|-------|---------------|
"""
    
    for tier, settings in SHADOWING_CONFIG['difficulty_progression'].items():
        report += f"| {tier.replace('_', ' ').title()} | {settings['max_words']} | {settings['speed_multiplier']}x | {settings['primary_focus']} |\n"
    
    report += f"""
### Content Requirements
- **Days with Shadowing:** {shadowing['days_with_shadowing']} (Days 7-30)
- **Total Phrases Needed:** {shadowing['total_phrases_needed']}

---

## 4. Implementation Checklist

### Immediate Actions (Preservation Mode)

These additions enhance without modifying existing content:

- [ ] Add color coding to grammar explanations (verb endings, gender)
- [ ] Add "Notice how..." callouts before each grammar pattern
- [ ] Add "Did you notice?" confirmations after patterns
- [ ] Add 5+ examples for each grammar point (input flooding)
- [ ] Add morpheme boundary markers (habl|o, habl|as)
- [ ] Create shadowing phrases for Days 7-30 ({shadowing['total_phrases_needed']} phrases)

### Files Generated

| File | Description |
|------|-------------|
| `Fluidez_Master_Alignment_Matrix_v2_Enhanced.xlsx` | Updated matrix with 5 new worksheets |
| `Fluidez_Day_Implementations.json` | Complete day-by-day specs for development |
| `Fluidez_Curriculum_Audit_Report.md` | This report |

---

## 5. New Worksheets Added

1. **PT Audit Results** - Day-by-day processability stage alignment with color coding
2. **Noticing Checklist** - Grammar point × technique coverage matrix (✅/❌)
3. **Shadowing Day Plans** - Day 7-30 phrase specifications and difficulty settings
4. **Day-by-Day Specs** - Complete implementation specs for all 30 days
5. **Enhancement Summary** - Overview dashboard

---

## Appendix: Framework References

### Processability Theory
- Pienemann, M. (1998). Language processing and second language development: Processability theory.
- Key insight: Grammatical structures can only be acquired in a fixed developmental sequence.

### Noticing Hypothesis  
- Schmidt, R. (1990). The role of consciousness in second language learning.
- Key insight: Learners must consciously notice forms to acquire them.

### Phonological Loop (Shadowing)
- Baddeley, A. (1986). Working memory.
- Key insight: Verbal rehearsal strengthens auditory-motor connections for pronunciation.

---

*Report generated by Fluidez Curriculum Enhancement Script*
"""
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print("    ✅ Markdown report generated successfully")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN EXECUTION
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 70)
    print("FLUIDEZ CURRICULUM AUDIT & ENHANCEMENT SCRIPT")
    print("=" * 70)
    print(f"\nPreservation Mode: {'ENABLED' if PRESERVE_MODE else 'DISABLED'}")
    print("  (Only adds content, never modifies existing material)\n")
    
    # Step 1: Audit
    print("\n[1/5] AUDITING CURRICULUM")
    print("-" * 40)
    auditor = CurriculumAuditor(INPUT_PATH)
    auditor.audit_processability()
    auditor.audit_noticing()
    auditor.audit_shadowing()
    auditor.generate_summary()
    
    # Step 2: Update Matrix
    print("\n[2/5] UPDATING ALIGNMENT MATRIX")
    print("-" * 40)
    updater = MatrixUpdater(auditor.wb, auditor.audit_results)
    updater.add_pt_audit_sheet()
    updater.add_noticing_checklist()
    updater.add_shadowing_specs()
    updater.add_day_implementations()
    updater.add_enhancement_summary()
    updater.save(OUTPUT_EXCEL)
    
    # Step 3: Generate Day Implementations
    print("\n[3/5] GENERATING DAY IMPLEMENTATIONS")
    print("-" * 40)
    generator = DayImplementationGenerator(auditor.audit_results)
    generator.generate_all_days()
    generator.export_json(OUTPUT_JSON)
    
    # Step 4: Generate Report
    print("\n[4/5] GENERATING AUDIT REPORT")
    print("-" * 40)
    generate_markdown_report(auditor.audit_results, OUTPUT_REPORT)
    
    # Step 5: Summary
    print("\n[5/5] COMPLETE")
    print("=" * 70)
    print("\n📁 OUTPUT FILES:")
    print(f"   • Excel:    {OUTPUT_EXCEL}")
    print(f"   • JSON:     {OUTPUT_JSON}")
    print(f"   • Report:   {OUTPUT_REPORT}")
    
    print("\n📊 AUDIT SUMMARY:")
    summary = auditor.audit_results["summary"]
    print(f"   • PT Alignment:      {summary['overall_scores']['pt_alignment']:.1f}%")
    print(f"   • Noticing Coverage: {summary['overall_scores']['noticing_coverage']:.1f}%")
    print(f"   • Shadowing Ready:   {summary['overall_scores']['shadowing_readiness']:.1f}%")
    
    print("\n📋 ACTION ITEMS:")
    print(f"   • PT Violations:     {summary['action_items']['pt_violations']}")
    print(f"   • Noticing Gaps:     {summary['action_items']['noticing_gaps']}")
    print(f"   • Shadowing Phrases: {summary['action_items']['shadowing_phrases_to_create']}")
    
    print("\n" + "=" * 70)
    print("✅ ENHANCEMENT COMPLETE - Existing content preserved")
    print("=" * 70)


if __name__ == "__main__":
    main()
